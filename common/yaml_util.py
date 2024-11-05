import yaml
from string import Template
from openpyxl import load_workbook
import xlrd
from xlutils.copy import copy


class YamlUtil():

    def __init__(self,file):
        self.file = file

    def read_extract_yaml(self):
        """读取extract.yaml文件"""
        with open(self.file,mode='r',encoding='utf-8') as f:
            value =  yaml.load(stream=f,Loader=yaml.FullLoader)
            return value

    def write_extract_yaml(self,data):
        """写入extract.yaml文件"""
        with open(self.file,mode='a',encoding='utf-8')as f:
            yaml.dump(data=data,stream=f,allow_unicode=True)

    def clear_extract_yaml(self):
        """清除extract.yaml文件"""
        with open(self.file,mode='w',encoding='utf-8') as f:
            f.truncate()

    def read_testcase_yaml(self):
        """读取测试用利的yaml文件"""
        with open(self.file,mode='r',encoding='utf-8') as f:
            value = yaml.load(stream=f,Loader=yaml.FullLoader)
            return value


    def read_excel(self,excel_file,sheet_name):
        """获取excel文件方法"""
        wb = load_workbook(excel_file)  # 打开文件
        sheet = wb[sheet_name]  # 表单定位
        test_data = []  # 创建一个空列表备用
        # 用for循环遍历表单，取到所有行的数据
        # i 代表行数，2表示从第二行开始读(一般情况下第一行是表头)，
        # sheet.max_ro+1表示读到最大行，并且读到最大行那行数据
        for i in range(2, sheet.max_row + 1):  # 2是起始值，sheet.max_row+1是中止值
            sub_data = {}  # 创建一个空字典，用来接收读取到的数据
            # 字典key值不同则增加元素，这个key和表头保持一致，i是行数，1是实际列数
            sub_data['case_id'] = sheet.cell(i, 1).value  # 将case_id这个key追加到sub_data字典，他的值是这个表单的第i行第一列
            sub_data['url'] = sheet.cell(i, 4).value
            sub_data['data'] = sheet.cell(i, 5).value
            sub_data['is_positive'] = sheet.cell(i, 6).value
            sub_data['request_type'] = sheet.cell(i, 8).value
            sub_data['expected_result'] = sheet.cell(i,9).value
            sub_data['status_code'] = sheet.cell(i,10).value
            sub_data['field_name'] = sheet.cell(i,12).value
            test_data.append(sub_data)
        return test_data

    def read_yaml_excel(self,excel_file,sheet_name):
        """读取excel文件至yaml"""
        value = []
        for i in self.read_excel(excel_file,sheet_name):
            with open(self.file,encoding='utf-8') as f:
                re = Template(f.read()).substitute(i)
                ls = yaml.safe_load(stream=re)
            for j in ls:
                value.append(j)
        return value

    def write_mysql_excel_yaml(self,mysql,sheet_name,excel):
        """同时读取yaml以及数据库文件方法"""
        for w in mysql.values():
            with open(self.file, encoding='utf-8') as f:
                self.l = w
        for i in self.read_excel('./test_excel/datas.xlsx',sheet_name):
            self.values = i[excel]
        with open(self.file,encoding='utf-8') as f:
            re = Template(f.read()).safe_substitute(datas1=self.l,datas2 = self.values)
            ls = yaml.safe_load(stream=re)
        return ls

    def datas_yaml(self,sheet_name,datas):
        """读取excel文件至yaml"""
        for i in self.read_excel('../test_excel/datas.xlsx',sheet_name):
            with open(self.file,encoding='utf-8') as f:
                value = i[datas]
                re = Template(f.read()).safe_substitute(datas = value)
        return re

    def write_excel_xls_append(path, value):
        """excel写入方法"""
        index = len(value)  # 获取需要写入数据的行数
        workbook = xlrd.open_workbook(path)  # 打开工作簿
        sheets = workbook.sheet_names()  # 获取工作簿中的所有表格
        worksheet = workbook.sheet_by_name(sheets[0])  # 获取工作簿中所有表格中的的第一个表格
        rows_old = worksheet.nrows  # 获取表格中已存在的数据的行数
        new_workbook = copy(workbook)  # 将xlrd对象拷贝转化为xlwt对象
        new_worksheet = new_workbook.get_sheet(0)  # 获取转化后工作簿中的第一个表格
        for i in range(0, index):
            for j in range(0, len(value[i])):
                new_worksheet.write(i + rows_old, j, value[i][j])  # 追加写入数据，注意是从i+rows_old行开始写入
        new_workbook.save(path)  # 保存工作簿


